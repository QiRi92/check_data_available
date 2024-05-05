# -*- coding: utf-8 -*-
"""
Created on Thu Mar 12 14:31:38 2020

@author: MY_EMEDUser07
"""

import os
import pandas as pd
import sys
import time
from openpyxl import load_workbook

from dx.lib.dxfile import Dxfile
from dx.lib.workfile import Workfile, WfTable, WfSeries
from dx.lib.series import Series

#FolderPath
FolderPath_Output = r"M:\Project\NewProject\03_Output"
FolderPath_LatestDxx = r"M:\LatestdXX"

#List
#List_Sectors = ['G7','Americas','Asia','Europe','CIS']
List_Sectors = ['Europe']
List_Column_Name = ["Sector", "Country_Code", "Section", "Table_Name", "dxCode", "dxDescription", "Frequency", "Sample", "Unit"]

def save_excel(df, sheetname):
    """This function is to save data into specific sheet name
    """
    wb = load_workbook(os.path.join(FolderPath_Output, sheetname + ".xlsx"))
    writer = pd.ExcelWriter(os.path.join(FolderPath_Output, sheetname + ".xlsx"), engine = "openpyxl")
    writer.book = wb
    #Delete the existing sheet which contains old data
    if sheetname in wb.sheetnames:
        del wb[sheetname]
    writer.sheets = dict((ws.title,ws) for ws in wb.worksheets)
    #Save the data into specific sheetname
    df.to_excel(writer, sheet_name = sheetname, index = False)
    #save the workbook
    writer.save()

def print_series(sector, folder, dxxfile, df):
    
    read_dxx_file = Workfile(os.path.join(folder, dxxfile))
    
    for table in read_dxx_file.tables:
        if "discontinued" in table.description and "Discontinued" in table.description:
            break
        else:
            for source_series in table:
                if "discontinued" in source_series.series.description or "Discontinued" in source_series.series.description:
                    break
                else:
                    if str(source_series.series.identifier) != "":
                        new_row = {}
                        new_row["Sector"] = sector
                        new_row["Country_Code"] = dxxfile[:2].upper()
                        new_row["Section"] = dxxfile[3:4].upper()
                        new_row["Table_Name"] = table.description
                        new_row["dxCode"] = source_series.series.identifier
                        new_row["dxDescription"] = source_series.series.description
                        new_row["Frequency"] = source_series.series.frequency
                        new_row["Sample"] = source_series.series.sample
                        new_row["Unit"] = source_series.series.units
                        df = df.append(new_row, ignore_index = True)
    return df
    
def main():
    
    for sector in List_Sectors:
        
        df = pd.DataFrame(columns = List_Column_Name)
        folder = os.path.join(FolderPath_LatestDxx, sector)
        
        for dxxfile in os.listdir(folder):
            if dxxfile.endswith(".dxx") and len(dxxfile) == 8:
                print(dxxfile)
                df = print_series(sector, folder, dxxfile, df)
        
        df.to_excel(os.path.join(FolderPath_Output, sector + ".xlsx"), index = False, engine='xlsxwriter')
        
if __name__ == "__main__":

    main()
