# -*- coding: utf-8 -*-
"""
Created on Fri Mar 13 09:41:21 2020

@author: MY_EMEDUser07
"""

import os
import pandas as pd
from openpyxl import load_workbook


Path_Source = r"M:\Project\NewProject\03_Output"

File_Information = r"M:\Project\NewProject\04_Information\Indicator_Information.xlsx"
File_Country = r"M:\Documentation\Coding\Country Code.xlsx"

List_Column_Name = ["Indicator", "Sector", "Country_Code", "Country_Name", "Section", "Table_Name", "dxCode", "dxDescription", "Frequency"]

def save_excel(df, sheetname):
    """This function is to save data into specific sheet name
    """
    wb = load_workbook(os.path.join(Path_Source, "result.xlsx"))
    writer = pd.ExcelWriter(os.path.join(Path_Source, "result.xlsx"), engine = "openpyxl")
    writer.book = wb
    #Delete the existing sheet which contains old data
    if sheetname in wb.sheetnames:
        del wb[sheetname]
    writer.sheets = dict((ws.title,ws) for ws in wb.worksheets)
    #Save the data into specific sheetname
    if sheetname == "Summary" or sheetname == "Summary_Freq":
        df.to_excel(writer, sheet_name = sheetname, index = True)
    else:
        df.to_excel(writer, sheet_name = sheetname, index = False)
    #save the workbook
    writer.save()
    
def summary_result(df):
    
    column_name = df[["Country_Name", "Country_Code"]].drop_duplicates().sort_values(by = ["Country_Name"])
    columns_index = pd.MultiIndex.from_frame(column_name)
    
    frequency_list = ["Daily", "Monthly", "Quarterly", "SemiAnnually", "Annually"]
    columns_name = []
    for row, column in column_name.iterrows():
        for x in frequency_list:
            c = (column["Country_Name"], column["Country_Code"], x)
            columns_name.append(c)
    columns_freq = pd.MultiIndex.from_tuples(columns_name)
    
    list_sector = df["Sector"].unique().tolist()
    
    df_summary = pd.DataFrame("N", columns = columns_index, index = pd.Index(df["Indicator"].unique().tolist()))
    df_summary_frequency = pd.DataFrame("N", columns = columns_freq, index = pd.Index(df["Indicator"].unique().tolist()))
    
    for index, col in df.iterrows():
        country = col["Country_Code"]
        countryname = col["Country_Name"]
        indicator = col["Indicator"]
        freq = col["Frequency"]
        df_summary.loc[indicator, (countryname, country)] = "Y"
        
        if freq == 1 or freq == 2:
            df_summary_frequency.loc[indicator, (countryname, country, "Annually")] = "Y"
        elif freq == 3:
            df_summary_frequency.loc[indicator, (countryname, country, "SemiAnnually")] = "Y"
        elif freq == 4:
            df_summary_frequency.loc[indicator, (countryname, country, "Quarterly")] = "Y"
        elif freq == 5:
            df_summary_frequency.loc[indicator, (countryname, country, "Monthly")] = "Y"
        elif freq == 7 or freq == 8:
            df_summary_frequency.loc[indicator, (countryname, country, "Daily")] = "Y"
        
    
    # df_summary = df_summary.sort_index(axis=1)
    
    save_excel(df_summary, "Summary")
    save_excel(df_summary_frequency, "Summary_Freq")
    #df_summary.to_excel(os.path.join(Path_Source, "result.xlsx"), sheet_name = "Summary")
    
    for sector in list_sector:
        df_sector = df.loc[df["Sector"] == sector]
        
        save_excel(df_sector, sector)
    
def search_key(indicator, List_key, df_result):
    
    for file in os.listdir(Path_Source):
        print(file)
        df_source = pd.read_excel(os.path.join(Path_Source,file), keep_default_na=False)
        df_source["Indicator"] = indicator
        

        df_filter1 = df_source.loc[df_source["Table_Name"].str.contains('|'.join(List_key), case = False)]
        df_filter2 = df_source.loc[df_source["dxDescription"].str.contains('|'.join(List_key), case = False)]
        
        df_result = df_result.append(df_filter1, ignore_index=True)
        df_result = df_result.append(df_filter2, ignore_index=True)
        df_result = df_result.drop_duplicates()
        
    return df_result

def main():
    
    # df_info = pd.read_excel(File_Information, sheet_name = "test")
    # #df_country = pd.read_excel(File_Country, sheet_name = "Country_List")
    # df_result = pd.DataFrame(columns = List_Column_Name)
    
    # for index, col in df_info.iterrows():
    #     indicator = col["Indicator"]
    #     List_key = col["Key_TableSeriesName"].split(",")
    #     print(indicator)
    #     df_result = search_key(indicator, List_key, df_result)
        
    # df_result.to_excel(os.path.join(Path_Source, "result.xlsx"), sheet_name = "SeriesFound", index = False)
    
    df_result = pd.read_excel(os.path.join(Path_Source, "result.xlsx"), sheet_name = "SeriesFound", keep_default_na = False)
    summary_result(df_result)
    
if __name__ == "__main__":

    main()